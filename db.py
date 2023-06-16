import psycopg2
import os
from dotenv import load_dotenv
from load_excel import *
from calculations import *
import openpyxl
import pandas as pd
from sqlalchemy import create_engine

load_dotenv()


class FirstDb:
    def __init__(self):
        """initialization of connection with db"""
        self.connection = psycopg2.connect(
            host=os.getenv("HOST"),
            user=os.getenv("USER"),
            password=os.getenv("PASSWORD"),
            database=os.getenv("DB_NAME"),
        )
        self.create_first_table()
        self.create_second_table()
        self.create_admin_table()
        self.create_patients_login_table()
        self.create_admin_full_table()
        self.create_users_full_table()

    def create_first_table(self):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """CREATE TABLE IF NOT EXISTS rand (
                amount INTEGER,
                measure TEXT,
                arithmetic double precision,
                minimal double precision,
                maximal double precision,
                deviation double precision,
                variation double precision,
                error double precision);"""
            )
        return "Таблиця готова до роботи"

    def get_user_info(self, login):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """SELECT * FROM full_info_users WHERE login = %s;""",
                (login,),
            )
            result = cursor.fetchall()
        return result

    def create_admin_full_table(self):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """CREATE TABLE IF NOT EXISTS full_info_excel (
                name TEXT,
                Age integer,
                AP integer,
                Height double precision,
                Weight double precision,
                BloodGroup integer,
                BloodSugar double precision,
                ATS integer,
                ATD integer,
                PAT integer,
                AtAverage double precision,
                Breathlessness integer,
                IndKetle double precision,
                IndKerdo double precision);"""
            )
        return "Таблиця готова до роботи"

    def insert_user(
        self,
        name,
        age,
        ap,
        height,
        weight,
        bloodgroup,
        bloodsugar,
        ats,
        atd,
        pat,
        ataverage,
        breathlessness,
        indketle,
        indkerdo,
    ):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """INSERT INTO full_info_excel (name, age, ap, height, weight, bloodgroup, bloodsugar, ats, atd, pat, ataverage, breathlessness, indketle, indkerdo) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (
                    name,
                    age,
                    ap,
                    height,
                    weight,
                    bloodgroup,
                    bloodsugar,
                    ats,
                    atd,
                    pat,
                    ataverage,
                    breathlessness,
                    indketle,
                    indkerdo,
                ),
            )
            self.connection.commit()

    def create_users_full_table(self):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """CREATE TABLE IF NOT EXISTS full_info_users (
                name TEXT,
                Age integer,
                AP integer,
                Height double precision,
                Weight double precision,
                BloodGroup integer,
                BloodSugar double precision,
                ATS integer,
                ATD integer,
                PAT integer,
                AtAverage double precision,
                Breathlessness integer,
                IndKetle double precision,
                IndKerdo double precision,
                login text);"""
            )
        return "Таблиця готова до роботи"

    def insert_full_excel(self, file_path):
        engine = create_engine(
            f'postgresql+psycopg2://{os.getenv("USER")}:{os.getenv("PASSWORD")}@{os.getenv("HOST")}/{os.getenv("DB_NAME")}'
        )
        with pd.ExcelFile(f"{file_path}") as xlsx:
            df = pd.read_excel(xlsx)
            df.to_sql(
                name="full_info_excel", con=engine, if_exists="append", index=False
            )

    def create_admin_table(self):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """CREATE TABLE IF NOT EXISTS users (
                login varchar,
                password varchar);"""
            )
        return "Таблиця готова до роботи"

    def create_patients_login_table(self):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """CREATE TABLE IF NOT EXISTS users_pat (
                login varchar,
                password varchar);"""
            )
        return "Таблиця готова до роботи"

    def get_list_of_colnames(self):
        with self.connection.cursor() as cursor:
            cursor.execute("""SELECT * FROM rand""")
            colnames = [desc[0] for desc in cursor.description]
        print(colnames)

    def check_if_excists_user(self, login):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """SELECT login FROM full_info_users WHERE login = %s;""",
                (login,),
            )
            result = cursor.fetchall()
        return bool(len(result))

    def check_if_exists_measure(self, measure):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """SELECT measure FROM rand WHERE measure = %s;""",
                (measure,),
            )
            result = cursor.fetchall()
        return bool(len(result))

    def check_if_excists_login(self, login):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """SELECT login FROM users_pat WHERE login = %s;""",
                (login,),
            )
            result = cursor.fetchall()
        return bool(len(result))

    def check_if_excists_password(self, password):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """SELECT password FROM users_pat WHERE password = %s;""",
                (password,),
            )
            result = cursor.fetchall()
        return bool(len(result))

    def check_if_user_data_exists(self, login):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """SELECT login FROM full_info_users WHERE login = %s;""",
                (login,),
            )
            result = cursor.fetchall()
        return bool(len(result))

    def check_if_exists_login_adm(self, login):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """SELECT login FROM users WHERE login = %s;""",
                (login,),
            )
            result = cursor.fetchall()
        return bool(len(result))

    def check_if_exists_password_adm(self, password):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """SELECT password FROM users WHERE password = %s;""",
                (password,),
            )
            result = cursor.fetchall()
        return bool(len(result))

    def check_if_exists_measure1(self, measure):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """SELECT measure FROM rand2 WHERE measure = %s;""",
                (measure,),
            )
            result = cursor.fetchall()
        return bool(len(result))

    def insert_new_user(self, login, password):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """INSERT INTO users_pat (login, password) VALUES (%s, %s)""",
                (
                    login,
                    password,
                ),
            )
            self.connection.commit()

    def insert_user_login(
        self,
        name,
        age,
        ap,
        height,
        weight,
        bloodgroup,
        bloodsugar,
        ats,
        atd,
        pat,
        ataverage,
        breathlessness,
        indketle,
        indkerdo,
        login,
    ):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """INSERT INTO full_info_users (name, age, ap, height, weight, bloodgroup, bloodsugar, ats, atd, pat, ataverage, breathlessness, indketle, indkerdo, login) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (
                    name,
                    age,
                    ap,
                    height,
                    weight,
                    bloodgroup,
                    bloodsugar,
                    ats,
                    atd,
                    pat,
                    ataverage,
                    breathlessness,
                    indketle,
                    indkerdo,
                    login,
                ),
            )
            self.connection.commit()

    def insert_arithmetic(self, amount, measure, value):
        if not self.check_if_exists_measure(measure):
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """INSERT INTO rand (amount, measure, arithmetic) VALUES (%s, %s, %s)""",
                    (
                        amount,
                        measure,
                        value,
                    ),
                )
                self.connection.commit()
        else:
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """UPDATE rand SET amount = %s, arithmetic = %s WHERE measure = %s;""",
                    (
                        amount,
                        value,
                        measure,
                    ),
                )
                self.connection.commit()

    def insert_minimal(self, amount, measure, value):
        if not self.check_if_exists_measure(measure):
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """INSERT INTO rand (amount, measure, minimal) VALUES (%s, %s, %s)""",
                    (
                        amount,
                        measure,
                        value,
                    ),
                )
                self.connection.commit()
        else:
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """UPDATE rand SET amount = %s, minimal = %s WHERE measure = %s;""",
                    (
                        amount,
                        value,
                        measure,
                    ),
                )
                self.connection.commit()

    def insert_maximal(self, amount, measure, value):
        if not self.check_if_exists_measure(measure):
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """INSERT INTO rand (amount, measure, maximal) VALUES (%s, %s, %s)""",
                    (
                        amount,
                        measure,
                        value,
                    ),
                )
                self.connection.commit()
        else:
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """UPDATE rand SET amount = %s, maximal = %s WHERE measure = %s;""",
                    (
                        amount,
                        value,
                        measure,
                    ),
                )
                self.connection.commit()

    def insert_deviation(self, amount, measure, value):
        if not self.check_if_exists_measure(measure):
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """INSERT INTO rand (amount, measure, deviation) VALUES (%s, %s, %s)""",
                    (
                        amount,
                        measure,
                        value,
                    ),
                )
                self.connection.commit()
        else:
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """UPDATE rand SET amount = %s, deviation = %s WHERE measure = %s;""",
                    (
                        amount,
                        value,
                        measure,
                    ),
                )
                self.connection.commit()

    def insert_variation(self, amount, measure, value):
        if not self.check_if_exists_measure(measure):
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """INSERT INTO rand (amount, measure, variation) VALUES (%s, %s, %s)""",
                    (
                        amount,
                        measure,
                        value,
                    ),
                )
                self.connection.commit()
        else:
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """UPDATE rand SET amount = %s, variation = %s WHERE measure = %s;""",
                    (
                        amount,
                        value,
                        measure,
                    ),
                )
                self.connection.commit()

    def insert_error(self, amount, measure, value):
        if not self.check_if_exists_measure(measure):
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """INSERT INTO rand (amount, measure, error) VALUES (%s, %s, %s)""",
                    (
                        amount,
                        measure,
                        value,
                    ),
                )
                self.connection.commit()
        else:
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """UPDATE rand SET amount = %s, error = %s WHERE measure = %s;""",
                    (
                        amount,
                        value,
                        measure,
                    ),
                )
                self.connection.commit()

    def create_second_table(self):
        with self.connection.cursor() as cursor:
            cursor.execute(
                """CREATE TABLE IF NOT EXISTS rand2 (
                amount INTEGER,
                measure TEXT,
                covariance double precision,
                pearson double precision,
                t_test varchar);"""
            )
        return "Таблиця готова до роботи"

    def insert_covariance(self, amount, measure, value):
        if not self.check_if_exists_measure1(measure):
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """INSERT INTO rand2 (amount, measure, covariance) VALUES (%s, %s, %s)""",
                    (
                        amount,
                        measure,
                        value,
                    ),
                )
                self.connection.commit()
        else:
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """UPDATE rand2 SET amount = %s, covariance = %s WHERE measure = %s;""",
                    (
                        amount,
                        value,
                        measure,
                    ),
                )
                self.connection.commit()

    def insert_pearson(self, amount, measure, value):
        if not self.check_if_exists_measure1(measure):
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """INSERT INTO rand2 (amount, measure, pearson) VALUES (%s, %s, %s)""",
                    (
                        amount,
                        measure,
                        value,
                    ),
                )
                self.connection.commit()
        else:
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """UPDATE rand2 SET amount = %s, pearson = %s WHERE measure = %s;""",
                    (
                        amount,
                        value,
                        measure,
                    ),
                )
                self.connection.commit()

    def insert_t_test(self, amount, measure, value):
        if not self.check_if_exists_measure1(measure):
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """INSERT INTO rand2 (amount, measure, t_test) VALUES (%s, %s, %s)""",
                    (
                        amount,
                        measure,
                        value,
                    ),
                )
                self.connection.commit()
        else:
            with self.connection.cursor() as cursor:
                cursor.execute(
                    """UPDATE rand2 SET amount = %s, t_test = %s WHERE measure = %s;""",
                    (
                        amount,
                        value,
                        measure,
                    ),
                )
                self.connection.commit()

    def get_rand_excel(self):
        with self.connection.cursor() as cursor:
            cursor.execute("""SELECT * FROM rand""")
            results = cursor.fetchall()
            book = openpyxl.Workbook()
            sheet = book.active
            k = 1
            for row in [
                "Кількість елементів",
                "Показник",
                "Середнє арифметичне",
                "Мінімальне значення",
                "Максимальне значення",
                "Середньоквадратичне відхилення по вибірці",
                "Коефіцієнт варіації",
                "Помилка середнього",
            ]:
                cell = sheet.cell(row=1, column=k)
                cell.value = row
                k += 1
            i = 1
            for row in results:
                i += 1
                j = 1
                for col in row:
                    cell = sheet.cell(row=i, column=j)
                    cell.value = col
                    j += 1
            book.save("rand.xlsx")

    def get_full_zvit(self):
        with self.connection.cursor() as cursor:
            cursor.execute("""SELECT * FROM full_info_excel""")
            results = cursor.fetchall()
            book = openpyxl.Workbook()
            sheet = book.active
            k = 1
            for row in [
                "name",
                "age",
                "ap",
                "height",
                "weight",
                "bloodgroup",
                "bloodsugar",
                "ats",
                "atd",
                "pat",
                "ataverage",
                "breathlessness",
                "indketle",
                "indkerdo",
            ]:
                cell = sheet.cell(row=1, column=k)
                cell.value = row
                k += 1
            i = 1
            for row in results:
                i += 1
                j = 1
                for col in row:
                    cell = sheet.cell(row=i, column=j)
                    cell.value = col
                    j += 1
            book.save("Повний звіт.xlsx")

    def get_rand2_excel(self):
        with self.connection.cursor() as cursor:
            cursor.execute("""SELECT * FROM rand2""")
            results = cursor.fetchall()
            book = openpyxl.Workbook()
            sheet = book.active
            k = 1
            for row in [
                "Кількість елементів",
                "Показник",
                "Коваріація",
                "Коефіцієнт кореляції Пірсона",
                "Т-критерий Стюдента",
            ]:
                cell = sheet.cell(row=1, column=k)
                cell.value = row
                k += 1
            i = 1
            for row in results:
                i += 1
                j = 1
                for col in row:
                    cell = sheet.cell(row=i, column=j)
                    cell.value = col
                    j += 1
            book.save("rand2.xlsx")

    def get_list_all(self, value):
        with self.connection.cursor() as cursor:
            cursor.execute("""SELECT * FROM full_info_excel;""")
            result = cursor.fetchall()
        return [i[value] for i in result]
